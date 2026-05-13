import csv
import io
from pathlib import Path
from typing import List, Dict
import logging

import pdfplumber

logger = logging.getLogger(__name__)

# Supported extensions
_SUPPORTED = {".pdf", ".csv", ".xlsx", ".xls"}


def load_pdfs(docs_folder: str) -> List[Dict]:
    """
    Recursively scan *docs_folder* for PDF / CSV / Excel files and extract
    their text content.

    Returns a list of dicts:
        {
            "source":  relative path from docs_folder (str),
            "path":    absolute path (str),
            "content": full extracted text (str),
        }
    Files whose text extraction fails are skipped with a warning.
    """
    docs_path = Path(docs_folder).resolve()
    if not docs_path.exists():
        raise FileNotFoundError(f"Docs folder not found: {docs_path}")

    all_files = [
        p for p in sorted(docs_path.rglob("*"))
        if p.is_file() and p.suffix.lower() in _SUPPORTED
    ]

    if not all_files:
        logger.warning("No supported documents found in '%s'.", docs_path)
        return []

    documents: List[Dict] = []
    for file_path in all_files:
        ext = file_path.suffix.lower()
        rel = str(file_path.relative_to(docs_path))
        try:
            if ext == ".pdf":
                text = _extract_pdf(file_path)
            elif ext == ".csv":
                text = _extract_csv(file_path)
            else:  # .xlsx / .xls
                text = _extract_excel(file_path)
        except Exception as exc:
            logger.warning("Skipping '%s' — extraction failed: %s", rel, exc)
            continue

        if not text.strip():
            logger.warning("Skipping '%s' — no extractable text.", rel)
            continue

        documents.append({"source": rel, "path": str(file_path), "content": text})
        logger.info("Loaded '%s' (%d chars).", rel, len(text))

    return documents


# ── Extractors ────────────────────────────────────────────────────────────────

def _extract_pdf(path: Path) -> str:
    pages: List[str] = []
    with pdfplumber.open(path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            page_text = page.extract_text()
            if page_text:
                pages.append(page_text)
            else:
                logger.debug("Page %d of '%s' yielded no text.", page_num, path.name)
    return "\n\n".join(pages)


def _extract_csv(path: Path) -> str:
    """Convert CSV rows to plain text: 'col1: val1 | col2: val2 ...' per row."""
    rows: List[str] = []
    # Try UTF-8 first, fall back to latin-1
    for encoding in ("utf-8-sig", "latin-1"):
        try:
            text = path.read_text(encoding=encoding)
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                line = " | ".join(f"{k}: {v}" for k, v in row.items() if v)
                if line:
                    rows.append(line)
            return "\n".join(rows)
        except UnicodeDecodeError:
            continue
    return ""


def _extract_excel(path: Path) -> str:
    """Convert all sheets of an Excel file to plain text."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required to read .xlsx files — pip install openpyxl")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sections: List[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: List[str] = []
        headers: List[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            cells = [str(c) if c is not None else "" for c in row]
            if i == 0:
                headers = cells
            else:
                if any(cells):
                    if headers:
                        line = " | ".join(
                            f"{h}: {v}" for h, v in zip(headers, cells) if v
                        )
                    else:
                        line = " | ".join(c for c in cells if c)
                    if line:
                        rows.append(line)
        if rows:
            sections.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(sections)
